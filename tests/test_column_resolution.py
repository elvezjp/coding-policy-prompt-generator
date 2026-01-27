from pathlib import Path

import pytest
from openpyxl import Workbook

from coding_policy_prompt_generator.excel_io import ColumnConfig, generate_prompts


def _make_workbook(path: Path, headers: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    ws.cell(row=2, column=1, value="N-001")
    ws.cell(row=2, column=2, value="概要")
    wb.save(path)


def test_relaxed_header_matching_allows_spaces(tmp_path: Path) -> None:
    input_path = tmp_path / "relaxed.xlsx"
    output_path = tmp_path / "relaxed_out.xlsx"

    # Headers include spaces; CLI-style names do not.
    headers = ["項 番", "概 要", "詳 細リンク"]
    _make_workbook(input_path, headers)

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column=None,
            link_column="詳細リンク",
        ),
        sheet_prefix="PROMPT_",
        dry_run=True,
    )

    assert plan.processed_rules == 1


def test_missing_required_column_reports_available_headers_and_suggestions(tmp_path: Path) -> None:
    input_path = tmp_path / "missing.xlsx"
    headers = ["項番", "概要", "詳細リンク"]
    _make_workbook(input_path, headers)

    with pytest.raises(ValueError) as excinfo:
        generate_prompts(
            input_path=input_path,
            output_path=tmp_path / "missing_out.xlsx",
            index_sheet="一覧",
            header_row=1,
            columns=ColumnConfig(
                id_column="項番",
                summary_column="概要",
                description_column=None,
                link_column="詳細りんく",  # typo
            ),
            sheet_prefix="PROMPT_",
            dry_run=True,
        )

    message = str(excinfo.value)
    assert "Available headers:" in message
    assert "詳細リンク" in message
    assert "Did you mean:" in message
    assert "Sheet: 一覧" in message
    assert "Header row: 1" in message


def test_missing_index_sheet_reports_available_sheets(tmp_path: Path) -> None:
    input_path = tmp_path / "missing_sheet.xlsx"
    wb = Workbook()
    wb.active.title = "一覧"
    wb.save(input_path)

    with pytest.raises(ValueError) as excinfo:
        generate_prompts(
            input_path=input_path,
            output_path=tmp_path / "missing_sheet_out.xlsx",
            index_sheet="存在しないシート",
            header_row=1,
            columns=ColumnConfig(
                id_column="項番",
                summary_column="概要",
                description_column=None,
                link_column=None,
            ),
            sheet_prefix="PROMPT_",
            dry_run=True,
        )

    message = str(excinfo.value)
    assert "Index sheet not found" in message
    assert "Available sheets:" in message
    assert "一覧" in message
