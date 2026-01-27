from pathlib import Path

from openpyxl import Workbook, load_workbook

from coding_policy_prompt_generator.excel_io import ColumnConfig, generate_prompts


def _make_index_wb(path: Path, rows: list[tuple[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.cell(row=1, column=1, value="項番")
    ws.cell(row=1, column=2, value="概要")
    ws.cell(row=1, column=3, value="説明")

    for i, (rule_id, summary) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=rule_id)
        ws.cell(row=i, column=2, value=summary)

    wb.save(path)


def test_collision_suffix_is_stable_across_reruns(tmp_path: Path) -> None:
    input_path = tmp_path / "collision_input.xlsx"
    first_output = tmp_path / "collision_out1.xlsx"
    second_output = tmp_path / "collision_out2.xlsx"

    # These collide after forbidden character replacement.
    rows = [("A/B", "s1"), ("A?B", "s2")]
    _make_index_wb(input_path, rows)

    plan1 = generate_prompts(
        input_path=input_path,
        output_path=first_output,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    wb1 = load_workbook(first_output)
    prompt_sheets_1 = sorted(name for name in wb1.sheetnames if name.startswith("PROMPT_"))
    assert prompt_sheets_1 == ["PROMPT_A_B", "PROMPT_A_B_2"]
    assert len(plan1.created_sheets) == 2

    # Re-run using the generated workbook as the next input.
    plan2 = generate_prompts(
        input_path=first_output,
        output_path=second_output,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    wb2 = load_workbook(second_output)
    prompt_sheets_2 = sorted(name for name in wb2.sheetnames if name.startswith("PROMPT_"))

    # No new prompt sheets should be created on re-run.
    assert prompt_sheets_2 == prompt_sheets_1
    assert plan2.created_sheets == []
    assert sorted(plan2.updated_sheets) == prompt_sheets_1


def test_sheet_name_strips_edge_apostrophes(tmp_path: Path) -> None:
    input_path = tmp_path / "apostrophe_input.xlsx"
    output_path = tmp_path / "apostrophe_output.xlsx"

    _make_index_wb(input_path, [("'abc'", "summary")])

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    wb = load_workbook(output_path)
    prompt_sheets = [name for name in wb.sheetnames if name.startswith("PROMPT_")]
    assert len(prompt_sheets) == 1
    name = prompt_sheets[0]
    assert not name.startswith("'")
    assert not name.endswith("'")
