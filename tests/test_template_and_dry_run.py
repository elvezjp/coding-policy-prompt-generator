from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from coding_policy_prompt_generator.excel_io import ColumnConfig, generate_prompts


def _make_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    headers = ["項番", "分類", "カテゴリ", "概要", "説明"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)

    ws.cell(row=2, column=1, value="R-1")
    ws.cell(row=2, column=2, value="設計")
    ws.cell(row=2, column=3, value="命名")
    ws.cell(row=2, column=4, value="概要")
    ws.cell(row=2, column=5, value="補足")

    wb.save(path)


def test_template_file_is_applied(tmp_path: Path) -> None:
    pytest.importorskip("jinja2")
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path = tmp_path / "template.j2"

    _make_input(input_path)
    template_path.write_text(
        "RULE={{ rule_id }}\nCLASS={{ classification }}\nCAT={{ category }}\nSUMMARY={{ summary }}",
        encoding="utf-8",
    )

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        template_path=template_path,
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "RULE=R-1" in prompt
    assert "CLASS=設計" in prompt
    assert "CAT=命名" in prompt


def test_dry_run_does_not_write_output(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "dry_run_output.xlsx"
    _make_input(input_path)

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=True,
    )

    assert plan.processed_rules == 1
    assert output_path.exists() is False
    assert any(action.kind == "create" for action in plan.actions)


def test_builtin_template_strict_with_project_context(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        strictness="strict",
        project_context="Java 17, Spring Boot 3.x",
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "## プロジェクト前提" in prompt
    assert "Java 17, Spring Boot 3.x" in prompt
    assert "疑わしい場合は違反（NG）として判定" in prompt
    assert "軽微な違反も見逃さない" in prompt


def test_builtin_template_lenient_omits_project_context(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        strictness="lenient",
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "## プロジェクト前提" not in prompt
    assert "明らかな違反のみNGとする" in prompt
    assert "疑わしい場合はOKとし、理由に懸念点を記載する" in prompt
    assert "軽微な違反も見逃さない" not in prompt


def test_project_context_empty_string_is_omitted(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        project_context="   ",
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "## プロジェクト前提" not in prompt


def test_builtin_template_contains_new_sections(tmp_path: Path) -> None:
    """新セクション（重大度、適用範囲、グレーゾーン例）がプレースホルダとして含まれる"""
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    # 新セクションの存在確認（マークダウン形式）
    assert "| 重大度 | （未指定） |" in prompt  # 規約概要テーブル内
    assert "## 適用範囲・例外" in prompt
    assert "## グレーゾーンの具体例" in prompt
    # プレースホルダとして「未記載」が含まれる
    assert "（未記載）" in prompt
    # 記入ガイドの一部が含まれる
    assert "必須（ビルドエラー）" in prompt  # 重大度のガイド
    assert "- **対象**: クラス名、インターフェース名" in prompt
    assert "// OK: PascalCase" in prompt
    assert "// NG: 先頭小文字" in prompt
    assert "`HTTPClient`" in prompt


def test_invalid_strictness_raises_error(tmp_path: Path) -> None:
    """strictness に不正な値を指定するとエラーが発生する"""
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    with pytest.raises(ValueError, match="Invalid strictness"):
        generate_prompts(
            input_path=input_path,
            output_path=output_path,
            index_sheet="一覧",
            header_row=1,
            columns=ColumnConfig(
                id_column="項番",
                summary_column="概要",
                description_column="説明",
                link_column="説明",
            ),
            sheet_prefix="PROMPT_",
            dry_run=False,
            strictness="invalid",
        )


def test_jinja_template_can_use_new_variables(tmp_path: Path) -> None:
    """Jinjaテンプレートで新変数（strictness, project_context）が利用可能"""
    pytest.importorskip("jinja2")
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path = tmp_path / "template.j2"

    _make_input(input_path)
    template_path.write_text(
        "STRICTNESS={{ strictness }}\nCONTEXT={{ project_context }}",
        encoding="utf-8",
    )

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        template_path=template_path,
        strictness="lenient",
        project_context="Java 21",
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "STRICTNESS=lenient" in prompt
    assert "CONTEXT=Java 21" in prompt


def test_jinja_old_template_without_new_variables(tmp_path: Path) -> None:
    """新変数を使用しない旧Jinjaテンプレートでも正常に動作する"""
    pytest.importorskip("jinja2")
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path = tmp_path / "template.j2"

    _make_input(input_path)
    # 旧形式: 新変数（strictness, project_context）を使用しない
    template_path.write_text(
        "RULE={{ rule_id }}\nSUMMARY={{ summary }}",
        encoding="utf-8",
    )

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        template_path=template_path,
        strictness="strict",
        project_context="Java 17",
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "RULE=R-1" in prompt
    assert "SUMMARY=概要" in prompt
    # 新変数は参照していないのでテンプレートに含まれない
    assert "strict" not in prompt
    assert "Java 17" not in prompt


def test_jinja_template_project_context_empty_string(tmp_path: Path) -> None:
    """project_context 未指定時は空文字が渡され、'None' を出力しない"""
    pytest.importorskip("jinja2")
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path = tmp_path / "template.j2"

    _make_input(input_path)
    template_path.write_text(
        "CONTEXT={{ project_context }}",
        encoding="utf-8",
    )

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
        template_path=template_path,
    )

    wb = load_workbook(output_path)
    prompt = wb["PROMPT_R-1"]["A1"].value
    assert isinstance(prompt, str)
    assert "CONTEXT=" in prompt
    assert "None" not in prompt


def test_jinja_template_missing_marker_warns(tmp_path: Path) -> None:
    """マーカーが含まれないテンプレートは警告される"""
    pytest.importorskip("jinja2")
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path = tmp_path / "template.j2"

    _make_input(input_path)
    template_path.write_text(
        "RULE={{ rule_id }}\nSUMMARY={{ summary }}",
        encoding="utf-8",
    )

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=True,
        template_path=template_path,
    )

    assert any("missing rule_id marker" in message for message in plan.warnings)
