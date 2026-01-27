from pathlib import Path

from openpyxl import Workbook, load_workbook

from coding_policy_prompt_generator.excel_io import ColumnConfig, generate_prompts


def _make_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "コーディング規約一覧"

    # Row 1-2: description rows (as in sample)
    ws.cell(row=1, column=1, value="説明")
    ws.cell(row=2, column=1, value="説明2")

    # Row 3: header
    headers = ["項番", "分類", "カテゴリ", "概要", "説明"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=name)

    # Row 4+: data
    ws.cell(row=4, column=1, value="N-001")
    ws.cell(row=4, column=4, value="テスト概要")
    ws.cell(row=4, column=5, value="補足")

    wb.save(path)


def test_generate_prompts_minimal(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="コーディング規約一覧",
        header_row=3,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    assert plan.processed_rules == 1
    assert output_path.exists()

    wb = load_workbook(output_path)
    assert "PROMPT_N-001" in wb.sheetnames

    index_ws = wb["コーディング規約一覧"]
    link_value = index_ws.cell(row=4, column=5).value
    assert isinstance(link_value, str)
    assert "HYPERLINK" in link_value

    prompt_value = wb["PROMPT_N-001"]["A1"].value
    assert isinstance(prompt_value, str)
    assert "N-001" in prompt_value


def test_hyperlink_escapes_apostrophe_in_sheet_name(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.cell(row=1, column=1, value="項番")
    ws.cell(row=1, column=2, value="概要")
    ws.cell(row=1, column=3, value="説明")
    ws.cell(row=2, column=1, value="O'Reilly-1")
    ws.cell(row=2, column=2, value="概要")
    wb.save(input_path)

    generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    wb_out = load_workbook(output_path)
    link_value = wb_out["一覧"].cell(row=2, column=3).value
    assert isinstance(link_value, str)
    assert "O''Reilly-1" in link_value
