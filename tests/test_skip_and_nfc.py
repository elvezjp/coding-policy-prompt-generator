from pathlib import Path
from unicodedata import normalize as unicode_normalize

from openpyxl import Workbook, load_workbook

from coding_policy_prompt_generator.excel_io import ColumnConfig, generate_prompts
from coding_policy_prompt_generator.cli import main


def _save_workbook(path: Path, workbook: Workbook) -> None:
    workbook.save(path)


def test_skip_conditions_record_warnings(tmp_path: Path) -> None:
    input_path = tmp_path / "skip_input.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"

    headers = ["項番", "概要", "説明"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)

    # Row 2: both empty -> ignored
    # Row 3: summary only -> skipped with warning
    ws.cell(row=3, column=2, value="概要のみ")
    # Row 4: id only -> skipped with warning
    ws.cell(row=4, column=1, value="IDのみ")
    # Row 5: valid
    ws.cell(row=5, column=1, value="N-100")
    ws.cell(row=5, column=2, value="有効な概要")

    _save_workbook(input_path, wb)

    plan = generate_prompts(
        input_path=input_path,
        output_path=tmp_path / "skip_output.xlsx",
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=True,
    )

    assert plan.processed_rules == 1
    assert [rec.row for rec in plan.skipped_rows] == [3, 4]
    assert len(plan.warnings) == 2
    assert "Row 3 skipped: missing rule_id" in plan.warnings
    assert "Row 4 skipped: missing summary" in plan.warnings


def test_nfc_normalization_for_sheet_and_headers(tmp_path: Path) -> None:
    input_path = tmp_path / "nfc_input.xlsx"
    output_path = tmp_path / "nfc_output.xlsx"

    index_sheet_nfc = "コーディング規約一覧"
    index_sheet_nfd = unicode_normalize("NFD", index_sheet_nfc)

    header_id_nfc = "項番"
    header_summary_nfc = "概要"
    header_link_nfc = "説明"

    header_id_nfd = unicode_normalize("NFD", header_id_nfc)
    header_summary_nfd = unicode_normalize("NFD", header_summary_nfc)
    header_link_nfd = unicode_normalize("NFD", header_link_nfc)

    wb = Workbook()
    ws = wb.active
    ws.title = index_sheet_nfd

    ws.cell(row=1, column=1, value=header_id_nfd)
    ws.cell(row=1, column=2, value=header_summary_nfd)
    ws.cell(row=1, column=3, value=header_link_nfd)

    ws.cell(row=2, column=1, value="N-001")
    ws.cell(row=2, column=2, value="NFCテスト")

    # Pre-create a detail sheet with NFD name; the implementation should update it.
    detail_nfc = "PROMPT_N-001"
    detail_nfd = unicode_normalize("NFD", detail_nfc)
    detail_ws = wb.create_sheet(title=detail_nfd)
    detail_ws["A1"].value = "規約ID: N-001"

    _save_workbook(input_path, wb)

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet=index_sheet_nfc,
        header_row=1,
        columns=ColumnConfig(
            id_column=header_id_nfc,
            summary_column=header_summary_nfc,
            description_column=header_link_nfc,
            link_column=header_link_nfc,
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    assert plan.processed_rules == 1
    assert plan.created_sheets == []
    assert len(plan.updated_sheets) == 1

    out_wb = load_workbook(output_path)
    assert unicode_normalize("NFC", index_sheet_nfc) in {
        unicode_normalize("NFC", name) for name in out_wb.sheetnames
    }
    assert unicode_normalize("NFC", detail_nfc) in {
        unicode_normalize("NFC", name) for name in out_wb.sheetnames
    }

    # Ensure no duplicate detail sheet was created.
    detail_nfc_count = sum(
        1 for name in out_wb.sheetnames if unicode_normalize("NFC", name) == detail_nfc
    )
    assert detail_nfc_count == 1


def test_legacy_marker_is_accepted_for_updates(tmp_path: Path) -> None:
    input_path = tmp_path / "legacy_marker_input.xlsx"
    output_path = tmp_path / "legacy_marker_output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.cell(row=1, column=1, value="項番")
    ws.cell(row=1, column=2, value="概要")
    ws.cell(row=1, column=3, value="説明")
    ws.cell(row=2, column=1, value="N-200")
    ws.cell(row=2, column=2, value="概要")

    legacy_sheet = wb.create_sheet(title="PROMPT_N-200")
    legacy_sheet["A1"].value = "【ルールID】\nN-200"

    _save_workbook(input_path, wb)

    plan = generate_prompts(
        input_path=input_path,
        output_path=output_path,
        index_sheet="一覧",
        header_row=1,
        columns=ColumnConfig(
            id_column="項番",
            summary_column="概要",
            description_column="説明",
            link_column="説明",
        ),
        sheet_prefix="PROMPT_",
        dry_run=False,
    )

    assert plan.processed_rules == 1
    assert plan.created_sheets == []
    assert plan.updated_sheets == ["PROMPT_N-200"]


def test_reject_same_input_output_path(tmp_path: Path) -> None:
    input_path = tmp_path / "same.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.cell(row=1, column=1, value="項番")
    ws.cell(row=1, column=2, value="概要")
    ws.cell(row=1, column=3, value="説明")
    wb.save(input_path)

    exit_code = main([str(input_path), "--output", str(input_path)])
    assert exit_code == 1
