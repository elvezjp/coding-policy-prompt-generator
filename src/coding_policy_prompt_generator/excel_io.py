from __future__ import annotations

import difflib
import importlib.resources
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from unicodedata import normalize as unicode_normalize

from jinja2 import Environment
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

FORBIDDEN_SHEET_CHARS = set("[]:*?/\\")

# AIオーディター形式の定数
AI_AUDITOR_DESCRIPTION = "このコーディング規約ファイルは、AIオーディターで読み込むためのファイルです。"
AI_AUDITOR_CREDIT = "このファイルはcoding-policy-prompt-generatorによって作成されました。"
AI_AUDITOR_REPO_URL = "https://github.com/elvezjp/coding-policy-prompt-generator"
AI_AUDITOR_HEADERS = ["項番", "分類", "カテゴリ", "概要", "説明"]
AI_AUDITOR_HEADER_ROW = 3
AI_AUDITOR_DATA_START_ROW = 4


@dataclass
class SkipRecord:
    row: int
    reason: str


@dataclass
class PlanAction:
    kind: str  # "create" | "update" | "skip"
    row: int
    rule_id: str
    sheet_name: Optional[str] = None
    reason: Optional[str] = None


@dataclass
class GenerationPlan:
    processed_rules: int = 0
    created_sheets: List[str] = field(default_factory=list)
    updated_sheets: List[str] = field(default_factory=list)
    skipped_rows: List[SkipRecord] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    actions: List[PlanAction] = field(default_factory=list)


@dataclass
class RuleData:
    rule_id: str
    summary: str
    description: str
    classification: str
    category: str
    sheet_name: str


@dataclass
class ColumnConfig:
    id_column: str
    summary_column: str
    description_column: Optional[str]
    link_column: Optional[str]


@dataclass
class ResolvedColumns:
    id_col: int
    summary_col: int
    description_col: Optional[int]
    link_col: int
    classification_col: Optional[int]
    category_col: Optional[int]


def generate_prompts(
    input_path: Path,
    output_path: Path,
    *,
    index_sheet: Optional[str],
    header_row: int,
    columns: ColumnConfig,
    sheet_prefix: str,
    dry_run: bool,
    template_path: Optional[Path] = None,
    strictness: str = "strict",
    project_context: Optional[str] = None,
) -> GenerationPlan:
    normalized_strictness = strictness.strip().lower()
    if normalized_strictness not in {"strict", "lenient"}:
        raise ValueError(f"Invalid strictness: {strictness}. Use 'strict' or 'lenient'.")

    normalized_project_context = project_context
    if normalized_project_context is not None:
        normalized_project_context = normalized_project_context.strip()
        if not normalized_project_context:
            normalized_project_context = None

    workbook = load_workbook(filename=input_path)
    worksheet = _resolve_index_sheet(workbook, index_sheet)

    headers = _build_headers(worksheet, header_row)
    resolved = _resolve_columns(headers, columns)
    renderer = _build_renderer(template_path)

    plan = GenerationPlan()
    _process_rows(
        workbook=workbook,
        worksheet=worksheet,
        header_row=header_row,
        resolved=resolved,
        sheet_prefix=sheet_prefix,
        renderer=renderer,
        strictness=normalized_strictness,
        project_context=normalized_project_context,
        plan=plan,
    )

    if not dry_run:
        workbook.save(output_path)

    return plan


def _resolve_index_sheet(workbook: Workbook, index_sheet: Optional[str]) -> Worksheet:
    if index_sheet:
        requested = _nfc(index_sheet)
        nfc_to_actual = {_nfc(name): name for name in workbook.sheetnames}
        actual = nfc_to_actual.get(requested)
        if not actual:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"Index sheet not found: {index_sheet}. Available sheets: {available}")
        return workbook[actual]

    # Default: first sheet
    return workbook[workbook.sheetnames[0]]


@dataclass
class HeaderContext:
    sheet_name: str
    header_row: int
    # Exact (NFC) header name -> column index
    header_map: Dict[str, int]
    # Relaxed header key -> exact header names
    relaxed_to_exacts: Dict[str, List[str]]
    # Header display order for messaging
    headers_in_order: List[str]
    # Rightmost header column index
    rightmost_header_col: int


def _build_headers(worksheet: Worksheet, header_row: int) -> HeaderContext:
    if header_row < 1:
        raise ValueError("header_row must be >= 1")

    header_map: Dict[str, int] = {}
    relaxed_to_exacts: Dict[str, List[str]] = {}
    headers_in_order: List[str] = []
    rightmost_header_col = 0

    for cell in worksheet[header_row]:
        if cell.value is None:
            continue
        name = _nfc(str(cell.value).strip())
        if not name:
            continue
        header_map[name] = cell.column
        headers_in_order.append(name)
        rightmost_header_col = max(rightmost_header_col, cell.column)

        relaxed_key = _relaxed_key(name)
        relaxed_to_exacts.setdefault(relaxed_key, []).append(name)

    if rightmost_header_col == 0:
        raise ValueError(f"No headers found on row {header_row} in sheet: {worksheet.title}")

    return HeaderContext(
        sheet_name=worksheet.title,
        header_row=header_row,
        header_map=header_map,
        relaxed_to_exacts=relaxed_to_exacts,
        headers_in_order=headers_in_order,
        rightmost_header_col=rightmost_header_col,
    )


def _resolve_columns(
    headers: HeaderContext,
    columns: ColumnConfig,
) -> ResolvedColumns:
    id_col = _require_column(headers, columns.id_column, logical_name="ID")
    summary_col = _require_column(headers, columns.summary_column, logical_name="summary")

    if columns.link_column:
        link_col = _require_column(headers, columns.link_column, logical_name="link")
    else:
        if headers.rightmost_header_col <= 0:
            raise ValueError(
                f"Unable to resolve link column from header row {headers.header_row} in sheet: {headers.sheet_name}"
            )
        link_col = headers.rightmost_header_col

    description_col: Optional[int] = None
    if columns.description_column:
        description_col = _find_column(headers, columns.description_column)

    classification_col = _find_column(headers, "分類")
    category_col = _find_column(headers, "カテゴリ")

    # If description and link point to the same column, ignore description.
    if description_col == link_col:
        description_col = None

    return ResolvedColumns(
        id_col=id_col,
        summary_col=summary_col,
        description_col=description_col,
        link_col=link_col,
        classification_col=classification_col,
        category_col=category_col,
    )


def _require_column(headers: HeaderContext, name: str, logical_name: str) -> int:
    col = _find_column(headers, name, strict=True)
    if col:
        return col

    available = ", ".join(headers.headers_in_order)
    suggestions = _suggest_headers(headers, name)
    suggestion_text = f" Did you mean: {', '.join(suggestions)}?" if suggestions else ""
    raise ValueError(
        (
            f"Required {logical_name} column not found: {name}."
            f"{suggestion_text} Sheet: {headers.sheet_name}. Header row: {headers.header_row}. "
            f"Available headers: {available}"
        )
    )


def _find_column(headers: HeaderContext, name: str, *, strict: bool = False) -> Optional[int]:
    requested_nfc = _nfc(name)
    exact_col = headers.header_map.get(requested_nfc)
    if exact_col:
        return exact_col

    relaxed = _relaxed_key(requested_nfc)
    matches = headers.relaxed_to_exacts.get(relaxed, [])
    if not matches:
        return None
    if len(matches) > 1:
        if strict:
            raise ValueError(
                f"Ambiguous header '{name}' matched multiple columns: {', '.join(matches)}"
            )
        return None
    return headers.header_map.get(matches[0])


def _suggest_headers(headers: HeaderContext, name: str) -> List[str]:
    relaxed = _relaxed_key(_nfc(name))
    candidates = difflib.get_close_matches(
        relaxed, headers.relaxed_to_exacts.keys(), n=3, cutoff=0.5
    )
    if candidates:
        return [headers.relaxed_to_exacts[key][0] for key in candidates]

    # Fallback for languages/typos where get_close_matches is too strict.
    scored: List[Tuple[float, str]] = []
    for key in headers.relaxed_to_exacts.keys():
        ratio = difflib.SequenceMatcher(a=relaxed, b=key).ratio()
        scored.append((ratio, key))
    scored.sort(reverse=True)

    if not scored or scored[0][0] < 0.3:
        return []
    return [headers.relaxed_to_exacts[scored[0][1]][0]]


def _rebuild_index_sheet(worksheet: Worksheet, rules: List[RuleData]) -> None:
    """一覧シートをAIオーディター形式で再構築する。"""
    # 既存データをクリア
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
            cell.hyperlink = None

    # Row 1: 説明文
    worksheet.cell(row=1, column=1, value=AI_AUDITOR_DESCRIPTION)

    # Row 2: クレジット＋ハイパーリンク
    credit_cell = worksheet.cell(row=2, column=1, value=AI_AUDITOR_CREDIT)
    credit_cell.hyperlink = AI_AUDITOR_REPO_URL
    credit_cell.font = Font(color="0000FF", underline="single")

    # Row 3: ヘッダ行
    for col, header in enumerate(AI_AUDITOR_HEADERS, start=1):
        worksheet.cell(row=AI_AUDITOR_HEADER_ROW, column=col, value=header)

    # Row 4+: データ行
    for i, rule in enumerate(rules):
        row_idx = AI_AUDITOR_DATA_START_ROW + i
        worksheet.cell(row=row_idx, column=1, value=rule.rule_id)
        worksheet.cell(row=row_idx, column=2, value=rule.classification)
        worksheet.cell(row=row_idx, column=3, value=rule.category)
        worksheet.cell(row=row_idx, column=4, value=rule.summary)
        # 列5（説明）にハイパーリンクを設定（descriptionは詳細シートに記載されるため上書き）
        link_cell = worksheet.cell(row=row_idx, column=5, value=_hyperlink_formula(rule.sheet_name))
        link_cell.font = Font(color="0000FF", underline="single")


def _process_rows(
    *,
    workbook: Workbook,
    worksheet: Worksheet,
    header_row: int,
    resolved: ResolvedColumns,
    sheet_prefix: str,
    renderer: "PromptRenderer",
    strictness: str,
    project_context: Optional[str],
    plan: GenerationPlan,
) -> None:
    rules: List[RuleData] = []

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        rule_id_raw = worksheet.cell(row=row_idx, column=resolved.id_col).value
        summary_raw = worksheet.cell(row=row_idx, column=resolved.summary_col).value

        rule_id = _clean_cell(rule_id_raw)
        summary = _clean_cell(summary_raw)

        if not rule_id and not summary:
            continue
        if not rule_id and summary:
            plan.skipped_rows.append(SkipRecord(row=row_idx, reason="missing rule_id"))
            plan.warnings.append(f"Row {row_idx} skipped: missing rule_id")
            plan.actions.append(PlanAction(kind="skip", row=row_idx, rule_id="", reason="missing rule_id"))
            continue
        if rule_id and not summary:
            plan.skipped_rows.append(SkipRecord(row=row_idx, reason="missing summary"))
            plan.warnings.append(f"Row {row_idx} skipped: missing summary")
            plan.actions.append(PlanAction(kind="skip", row=row_idx, rule_id=rule_id, reason="missing summary"))
            continue

        description = ""
        if resolved.description_col:
            description = _clean_cell(worksheet.cell(row=row_idx, column=resolved.description_col).value)

        classification = ""
        if resolved.classification_col:
            classification = _clean_cell(worksheet.cell(row=row_idx, column=resolved.classification_col).value)

        category = ""
        if resolved.category_col:
            category = _clean_cell(worksheet.cell(row=row_idx, column=resolved.category_col).value)

        sheet_name, action_kind = _ensure_detail_sheet(workbook, sheet_prefix, rule_id, plan)
        prompt = renderer.render(
            rule_id=rule_id,
            summary=summary,
            description=description,
            classification=classification,
            category=category,
            strictness=strictness,
            project_context=project_context,
        )
        markers = _rule_id_markers(rule_id)
        if not any(marker in prompt for marker in markers):
            plan.warnings.append(
                (
                    f"Row {row_idx} warning: prompt missing rule_id marker; "
                    "idempotent updates may not work"
                )
            )

        detail_ws = workbook[sheet_name]
        detail_cell = detail_ws["A1"]
        detail_cell.value = prompt
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
        detail_ws.column_dimensions["A"].width = 80

        rules.append(RuleData(
            rule_id=rule_id,
            summary=summary,
            description=description,
            classification=classification,
            category=category,
            sheet_name=sheet_name,
        ))

        plan.processed_rules += 1
        plan.actions.append(PlanAction(kind=action_kind, row=row_idx, rule_id=rule_id, sheet_name=sheet_name))

    # 一覧シートをAIオーディター形式で再構築
    _rebuild_index_sheet(worksheet, rules)


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _ensure_detail_sheet(workbook: Workbook, sheet_prefix: str, rule_id: str, plan: GenerationPlan) -> Tuple[str, str]:
    base_name = f"{sheet_prefix}{rule_id}"
    normalized_base = _normalize_sheet_name(base_name)
    markers = _rule_id_markers(rule_id)

    existing_by_nfc = {_nfc(name): name for name in workbook.sheetnames}
    existing_name = existing_by_nfc.get(_nfc(normalized_base))
    if existing_name:
        if _sheet_matches_marker(workbook, existing_name, markers):
            plan.updated_sheets.append(existing_name)
            return existing_name, "update"
        plan.warnings.append(
            (
                f"Sheet name collision for rule_id={rule_id}: "
                f"existing sheet '{existing_name}' did not match marker; creating a new sheet."
            )
        )

    # Idempotency guard: if the detail sheet was previously created with a suffix
    # due to collisions, try to find it by rule_id marker in A1.
    matched = _find_existing_detail_sheet_by_marker(workbook, markers)
    if matched:
        plan.updated_sheets.append(matched)
        return matched, "update"

    unique_name = _make_unique_sheet_name(workbook, normalized_base)
    workbook.create_sheet(title=unique_name)
    plan.created_sheets.append(unique_name)
    return unique_name, "create"


def _normalize_sheet_name(name: str) -> str:
    nfc_name = _nfc(name)
    cleaned: List[str] = []
    for ch in nfc_name:
        cleaned.append("_" if ch in FORBIDDEN_SHEET_CHARS else ch)
    normalized = "".join(cleaned).strip()
    normalized = " ".join(normalized.split())
    # Excel does not allow sheet names that begin or end with an apostrophe.
    normalized = normalized.strip("'")
    if not normalized:
        normalized = "PROMPT"
    return normalized[:31]


def _make_unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    existing_nfc = {_nfc(name) for name in workbook.sheetnames}
    if _nfc(base_name) not in existing_nfc:
        return base_name

    counter = 2
    while True:
        suffix = f"_{counter}"
        max_base_len = 31 - len(suffix)
        candidate = f"{base_name[:max_base_len]}{suffix}"
        if _nfc(candidate) not in existing_nfc:
            return candidate
        counter += 1


def _nfc(text: str) -> str:
    return unicode_normalize("NFC", text)


def _relaxed_key(text: str) -> str:
    # NFC + casefold + remove common separators/spaces for resilient matching.
    nfc = _nfc(text)
    folded = nfc.casefold()
    separators = " \t\r\n_-ー−‐　"
    return "".join(ch for ch in folded if ch not in separators)


def _rule_id_markers(rule_id: str) -> List[str]:
    return [
        f"規約ID: {rule_id}",
        f"規約ID | `{rule_id}`",  # Markdown table format
        f"【ルールID】\n{rule_id}",
    ]


def _find_existing_detail_sheet_by_marker(workbook: Workbook, markers: List[str]) -> Optional[str]:
    for sheet_name in workbook.sheetnames:
        if _sheet_matches_marker(workbook, sheet_name, markers):
            return sheet_name
    return None


def _sheet_matches_marker(workbook: Workbook, sheet_name: str, markers: List[str]) -> bool:
    value = workbook[sheet_name]["A1"].value
    if isinstance(value, str):
        return any(marker in value for marker in markers)
    return False


class PromptRenderer:
    def render(
        self,
        *,
        rule_id: str,
        summary: str,
        description: str,
        classification: str,
        category: str,
        strictness: str,
        project_context: Optional[str],
    ) -> str:
        raise NotImplementedError


class BuiltinPromptRenderer(PromptRenderer):
    def __init__(self) -> None:
        env = Environment(autoescape=False, trim_blocks=True, lstrip_blocks=True)
        templates_package = "coding_policy_prompt_generator.templates"

        system_text = importlib.resources.files(templates_package).joinpath("system_prompt.j2").read_text(encoding="utf-8")
        user_text = importlib.resources.files(templates_package).joinpath("user_prompt.j2").read_text(encoding="utf-8")

        self._system_template = env.from_string(system_text)
        self._user_template = env.from_string(user_text)

    def render(
        self,
        *,
        rule_id: str,
        summary: str,
        description: str,
        classification: str,
        category: str,
        strictness: str,
        project_context: Optional[str],
    ) -> str:
        safe_project_context = project_context or ""
        context = {
            "rule_id": rule_id,
            "summary": summary,
            "description": description,
            "classification": classification,
            "category": category,
            "strictness": strictness,
            "project_context": safe_project_context,
        }
        system_prompt = self._system_template.render(**context)
        user_prompt = self._user_template.render(**context)
        return system_prompt + user_prompt


class JinjaPromptRenderer(PromptRenderer):
    def __init__(self, template_text: str) -> None:
        try:
            from jinja2 import Environment, StrictUndefined
        except Exception as exc:  # pragma: no cover - import-time failure
            raise RuntimeError("jinja2 is required for --template support") from exc

        env = Environment(undefined=StrictUndefined, autoescape=False, trim_blocks=True, lstrip_blocks=True)
        self._template = env.from_string(template_text)

    def render(
        self,
        *,
        rule_id: str,
        summary: str,
        description: str,
        classification: str,
        category: str,
        strictness: str,
        project_context: Optional[str],
    ) -> str:
        safe_project_context = project_context or ""
        return self._template.render(
            rule_id=rule_id,
            summary=summary,
            description=description,
            classification=classification,
            category=category,
            strictness=strictness,
            project_context=safe_project_context,
        )


def _build_renderer(template_path: Optional[Path]) -> PromptRenderer:
    if not template_path:
        return BuiltinPromptRenderer()

    if not template_path.exists():
        raise ValueError(f"Template file not found: {template_path}")
    template_text = template_path.read_text(encoding="utf-8")
    return JinjaPromptRenderer(template_text)


def _hyperlink_formula(sheet_name: str, label: str = "詳細") -> str:
    # Excel internal link to A1 of the detail sheet.
    safe_sheet_name = sheet_name.replace("'", "''")
    return f"=HYPERLINK(\"#'{safe_sheet_name}'!A1\",\"{label}\")"
